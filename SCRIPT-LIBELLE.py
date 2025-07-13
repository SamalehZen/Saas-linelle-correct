#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de Traitement des Libellés d'Articles
===========================================

Ce script traite les libellés d'articles selon des règles spécifiques :
1. Supprime les accents et caractères spéciaux (sauf virgules)
2. Remplace les points par des virgules dans les grammages/volumes
3. Recompose : MARQUE + PRODUIT + GRAMMAGE/VOLUME
4. Met tout en MAJUSCULES
5. Conserve les fractions telles quelles (pas de conversion décimale)

Auteur: Assistant Scout - Scrapybara
Date: Juillet 2025
"""

import re
import unicodedata
import csv
import os
from typing import List, Tuple

# Installation automatique d'openpyxl si nécessaire
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    print("⚠️  Module openpyxl non trouvé. Installation automatique...")
    os.system("pip install openpyxl")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        EXCEL_AVAILABLE = True
        print("✅ Module openpyxl installé avec succès !")
    except ImportError:
        print("❌ Impossible d'installer openpyxl. Seul le fichier TSV sera généré.")
        EXCEL_AVAILABLE = False


class TraitementLibelles:
    """Classe pour traiter les libellés d'articles selon des règles spécifiques."""
    
    def __init__(self):
        """Initialise le processeur avec les marques reconnues."""
        self.marques = ['CRF', 'CARF', 'CARREFOUR', 'PAPERMATE', 'PM', 'SHARPIE', 'ROTRING']
    
    def supprimer_accents_et_caracteres_speciaux(self, texte: str) -> str:
        """
        Supprime les accents et caractères spéciaux, garde seulement les virgules.
        
        Args:
            texte (str): Le texte à nettoyer
            
        Returns:
            str: Le texte nettoyé
        """
        # Normaliser les caractères Unicode (décomposer les accents)
        normalise = unicodedata.normalize('NFD', texte)
        # Supprimer les marques diacritiques (accents)
        sans_accents = ''.join(char for char in normalise if unicodedata.category(char) != 'Mn')
        # Garder seulement les lettres, chiffres, espaces, virgules et quelques caractères autorisés
        nettoye = re.sub(r'[^A-Za-z0-9\s,./X+-]', ' ', sans_accents)
        # Nettoyer les espaces multiples
        nettoye = re.sub(r'\s+', ' ', nettoye).strip()
        return nettoye
    
    def extraire_marque(self, texte: str) -> str:
        """
        Extrait la marque du libellé.
        
        Args:
            texte (str): Le texte à analyser
            
        Returns:
            str: La marque trouvée ou None
        """
        texte_majuscule = texte.upper()
        
        for marque in self.marques:
            if re.search(r'\b' + re.escape(marque) + r'\b', texte_majuscule):
                return marque
        return None
    
    def extraire_grammage_volume(self, texte: str) -> List[str]:
        """
        Extrait le grammage/volume du texte de manière très précise.
        
        Args:
            texte (str): Le texte à analyser
            
        Returns:
            List[str]: Liste des grammages/volumes trouvés
        """
        texte_majuscule = texte.upper()
        grammages_volumes = []
        
        # Pattern 1: Grammages/volumes avec décimales
        pattern1 = r'\b\d+[.,]\d+\s*(?:KG|G|L|ML|CL)\b'
        matches1 = re.findall(pattern1, texte_majuscule)
        for match in matches1:
            nettoye = re.sub(r'(\d+)\.(\d+)', r'\1,\2', match)
            grammages_volumes.append(nettoye)
        
        # Pattern 2: Format X multiplicateur avec grammage
        pattern2 = r'\b\d+X\d+[.,]?\d*\s*(?:KG|G|L|ML|CL)\b'
        matches2 = re.findall(pattern2, texte_majuscule)
        for match in matches2:
            nettoye = re.sub(r'(\d+)\.(\d+)', r'\1,\2', match)
            grammages_volumes.append(nettoye)
        
        # Pattern 3: Grammages/volumes simples (>=10 pour éviter faux positifs)
        pattern3 = r'\b(?:[1-9]\d{2,}|[1-9]\d)\s*(?:KG|G|L|ML|CL)\b'
        matches3 = re.findall(pattern3, texte_majuscule)
        for match in matches3:
            grammages_volumes.append(match)
        
        # Pattern 4: Petits grammages/volumes (1-9) avec validation de contexte
        pattern4 = r'(?:^|\s)([1-9])\s*(KG|G|L|ML|CL)\b'
        matches4 = re.findall(pattern4, texte_majuscule)
        for match in matches4:
            grammages_volumes.append(match[0] + match[1])
        
        # Pattern 5: Fractions avec unité
        pattern5 = r'\b\d+/\d+\s*(?:KG|G|L|ML|CL)\b'
        matches5 = re.findall(pattern5, texte_majuscule)
        grammages_volumes.extend(matches5)
        
        # Pattern 6: Fractions sans unité
        pattern6 = r'\b\d+/\d+\b'
        matches6 = re.findall(pattern6, texte_majuscule)
        for match in matches6:
            if not any(match in gv for gv in grammages_volumes):
                grammages_volumes.append(match)
        
        # Supprimer les doublons
        grammages_uniques = []
        for gv in grammages_volumes:
            if gv not in grammages_uniques:
                grammages_uniques.append(gv)
        
        return grammages_uniques
    
    def traiter_libelle(self, libelle_original: str) -> str:
        """
        Traite un libellé selon les règles définies.
        
        Args:
            libelle_original (str): Le libellé original à traiter
            
        Returns:
            str: Le libellé traité
        """
        # Étape 1: Nettoyer les caractères spéciaux
        nettoye = self.supprimer_accents_et_caracteres_speciaux(libelle_original)
        
        # Étape 2: Extraire la marque
        marque = self.extraire_marque(nettoye)
        
        # Étape 3: Extraire grammages/volumes
        grammages_volumes = self.extraire_grammage_volume(nettoye)
        
        # Étape 4: Construire le nom du produit (sans marque ni grammage/volume)
        nom_produit = nettoye
        
        # Supprimer la marque
        if marque:
            nom_produit = re.sub(r'\b' + re.escape(marque) + r'\b', '', nom_produit, flags=re.IGNORECASE)
        
        # Supprimer les grammages/volumes
        for gv in grammages_volumes:
            gv_avec_point = gv.replace(',', '.')
            nom_produit = re.sub(r'\b' + re.escape(gv) + r'\b', '', nom_produit, flags=re.IGNORECASE)
            nom_produit = re.sub(r'\b' + re.escape(gv_avec_point) + r'\b', '', nom_produit, flags=re.IGNORECASE)
        
        nom_produit = re.sub(r'\s+', ' ', nom_produit).strip()
        
        # Étape 5: Assembler le libellé final
        parties_finales = []
        
        if marque:
            parties_finales.append(marque)
        
        if nom_produit:
            parties_finales.append(nom_produit)
        
        parties_finales.extend(grammages_volumes)
        
        libelle_final = ' '.join(parties_finales).upper()
        libelle_final = re.sub(r'\s+', ' ', libelle_final).strip()
        
        return libelle_final


def lire_articles_depuis_fichier(chemin_fichier: str) -> List[str]:
    """
    Lit une liste d'articles depuis un fichier texte.
    
    Args:
        chemin_fichier (str): Chemin vers le fichier contenant les articles
        
    Returns:
        List[str]: Liste des articles
    """
    try:
        with open(chemin_fichier, 'r', encoding='utf-8') as fichier:
            articles = [ligne.strip() for ligne in fichier if ligne.strip()]
        return articles
    except FileNotFoundError:
        print(f"❌ Fichier non trouvé: {chemin_fichier}")
        return []
    except Exception as e:
        print(f"❌ Erreur lors de la lecture du fichier: {e}")
        return []


def sauvegarder_tsv(resultats: List[Tuple[str, str]], chemin_sortie: str) -> bool:
    """
    Sauvegarde les résultats dans un fichier TSV.
    
    Args:
        resultats (List[Tuple[str, str]]): Liste des tuples (original, traité)
        chemin_sortie (str): Chemin du fichier de sortie
        
    Returns:
        bool: True si succès, False sinon
    """
    try:
        with open(chemin_sortie, 'w', newline='', encoding='utf-8') as fichier_tsv:
            writer = csv.writer(fichier_tsv, delimiter='\t')
            writer.writerow(['Libellé Original', 'Libellé Corrigé'])
            writer.writerows(resultats)
        return True
    except Exception as e:
        print(f"❌ Erreur lors de la sauvegarde TSV: {e}")
        return False


def sauvegarder_excel(resultats: List[Tuple[str, str]], chemin_sortie: str) -> bool:
    """
    Sauvegarde les résultats dans un fichier Excel formaté.
    
    Args:
        resultats (List[Tuple[str, str]]): Liste des tuples (original, traité)
        chemin_sortie (str): Chemin du fichier de sortie
        
    Returns:
        bool: True si succès, False sinon
    """
    if not EXCEL_AVAILABLE:
        print("⚠️  Module openpyxl non disponible. Fichier Excel non créé.")
        return False
    
    try:
        # Préparer les données avec en-tête
        donnees = [['Libellé Original', 'Libellé Corrigé']] + resultats
        
        # Créer un nouveau classeur Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Articles Traités"
        
        # Définir les styles
        font_entete = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        remplissage_entete = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        alignement_entete = Alignment(horizontal='center', vertical='center')
        
        font_cellule = Font(name='Calibri', size=11)
        alignement_cellule = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        bordure = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Ajouter les données
        for index_ligne, donnees_ligne in enumerate(donnees, 1):
            for index_colonne, valeur_cellule in enumerate(donnees_ligne, 1):
                cellule = ws.cell(row=index_ligne, column=index_colonne, value=valeur_cellule)
                
                # Style pour l'en-tête
                if index_ligne == 1:
                    cellule.font = font_entete
                    cellule.fill = remplissage_entete
                    cellule.alignment = alignement_entete
                else:
                    cellule.font = font_cellule
                    cellule.alignment = alignement_cellule
                
                cellule.border = bordure
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 50  # Libellé Original
        ws.column_dimensions['B'].width = 50  # Libellé Corrigé
        
        # Figer les volets (première ligne)
        ws.freeze_panes = 'A2'
        
        # Ajouter un filtre automatique
        ws.auto_filter.ref = f"A1:B{len(donnees)}"
        
        # Sauvegarder le fichier
        wb.save(chemin_sortie)
        return True
    except Exception as e:
        print(f"❌ Erreur lors de la sauvegarde Excel: {e}")
        return False


def traiter_articles(articles: List[str], nom_fichier_sortie: str = "articles_traites") -> None:
    """
    Traite une liste d'articles et génère les fichiers de sortie.
    
    Args:
        articles (List[str]): Liste des articles à traiter
        nom_fichier_sortie (str): Nom de base pour les fichiers de sortie
    """
    if not articles:
        print("❌ Aucun article à traiter.")
        return
    
    print(f"🔄 Traitement de {len(articles)} articles...")
    
    # Initialiser le processeur
    processeur = TraitementLibelles()
    
    # Traiter tous les articles
    resultats = []
    for article in articles:
        article_nettoye = article.strip()
        if article_nettoye:
            article_traite = processeur.traiter_libelle(article_nettoye)
            resultats.append((article_nettoye, article_traite))
    
    # Sauvegarder en TSV
    chemin_tsv = f"{nom_fichier_sortie}.tsv"
    if sauvegarder_tsv(resultats, chemin_tsv):
        print(f"✅ Fichier TSV créé: {chemin_tsv}")
    
    # Sauvegarder en Excel
    chemin_excel = f"{nom_fichier_sortie}.xlsx"
    if sauvegarder_excel(resultats, chemin_excel):
        print(f"✅ Fichier Excel créé: {chemin_excel}")
    
    print(f"\n📊 Résumé du traitement:")
    print(f"   • Articles traités: {len(resultats)}")
    print(f"   • Fichiers générés: {chemin_tsv}" + (f", {chemin_excel}" if EXCEL_AVAILABLE else ""))
    
    # Afficher quelques exemples
    print(f"\n🔍 Aperçu des résultats:")
    print("=" * 80)
    for i, (original, traite) in enumerate(resultats[:5]):
        print(f"{i+1:2d}. {original}")
        print(f"    → {traite}")
        print()


def main():
    """Fonction principale avec exemples d'utilisation."""
    print("=" * 80)
    print("🏷️  SCRIPT DE TRAITEMENT DES LIBELLÉS D'ARTICLES")
    print("=" * 80)
    print()
    
    # Exemple 1: Articles dans le code
    print("📝 Exemple 1: Articles définis dans le code")
    articles_exemple = [
        "6X30G CHIPS LISSE NAT CRF CLAS",
        "1L PET PUR JUS POMME CRF EXTRA",
        "PAPERMATE 4 Magic+ effaceurs fins réécr",
        "Désodorisant 2.5ml 4scent",
        "5 BQ ALU 1,5L PROFONDE"
    ]
    
    traiter_articles(articles_exemple, "exemple_articles")
    
    print("\n" + "="*80)
    print("📁 Exemple 2: Articles depuis un fichier")
    print("Pour traiter un fichier, utilisez :")
    print("   articles = lire_articles_depuis_fichier('mon_fichier.txt')")
    print("   traiter_articles(articles, 'mon_resultat')")
    print("="*80)


# EXEMPLES D'UTILISATION POUR VOS FUTURES LISTES
def exemple_utilisation_fichier():
    """
    Exemple d'utilisation pour traiter un fichier d'articles.
    Décommentez et adaptez selon vos besoins.
    """
    # # Lire les articles depuis un fichier
    # articles = lire_articles_depuis_fichier('ma_liste_articles.txt')
    # 
    # # Traiter et sauvegarder
    # traiter_articles(articles, 'mes_articles_traites')


def exemple_utilisation_liste():
    """
    Exemple d'utilisation pour traiter une liste d'articles définie dans le code.
    """
    # Vos articles ici
    mes_articles = [
        "Votre article 1",
        "Votre article 2",
        "Votre article 3",
        # Ajoutez tous vos articles ici...
    ]
    
    # Traiter et sauvegarder
    traiter_articles(mes_articles, 'ma_nouvelle_liste')


if __name__ == "__main__":
    main()